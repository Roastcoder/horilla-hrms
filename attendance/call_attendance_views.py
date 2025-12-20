"""
Call-based Auto Attendance Views

Views for managing call logs, attendance configuration, and manual updates.
"""

import csv
import io
from datetime import date, datetime, timedelta
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.core.paginator import Paginator
from django.db.models import Q, Sum, Count
from django.http import JsonResponse, HttpResponse
from django.utils.translation import gettext_lazy as _
from django.views.decorators.http import require_http_methods
from django.core.exceptions import PermissionDenied
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import pandas as pd

from employee.models import Employee
from .call_attendance_models import (
    CallLog, CallAttendanceConfig, CallAttendance, CallAttendanceAudit
)
from .call_attendance_forms import (
    CallLogForm, CallAttendanceConfigForm, ManualAttendanceUpdateForm,
    CallLogBulkUploadForm, AttendanceFilterForm
)
from .call_attendance_services import CallAttendanceService, CallLogService
from sales_incentive.validators import (
    ComprehensiveValidator, SalesValidationError, AttendanceValidator,
    IncentiveValidator, LoanPointValidator, AuditValidator
)


@login_required
def call_attendance_dashboard(request):
    """Dashboard view for call attendance overview"""
    today = date.today()
    current_month_start = date(today.year, today.month, 1)
    
    # Get statistics
    total_employees = Employee.objects.filter(is_active=True).count()
    today_attendances = CallAttendance.objects.filter(attendance_date=today)
    
    stats = {
        'total_employees': total_employees,
        'today_present': today_attendances.filter(attendance_status='PRESENT').count(),
        'today_half_day': today_attendances.filter(attendance_status='HALF_DAY').count(),
        'today_absent': today_attendances.filter(attendance_status='ABSENT').count(),
        'manual_updates_today': today_attendances.filter(source='MANUAL').count(),
    }
    
    # Recent attendance records
    recent_attendances = CallAttendance.objects.select_related(
        'employee_id'
    ).order_by('-attendance_date', '-updated_at')[:10]
    
    # Recent audit logs
    recent_audits = CallAttendanceAudit.objects.select_related(
        'employee_id', 'updated_by'
    ).order_by('-timestamp')[:5]
    
    context = {
        'stats': stats,
        'recent_attendances': recent_attendances,
        'recent_audits': recent_audits,
        'active_config': CallAttendanceService.get_active_config(),
    }
    
    return render(request, 'call_attendance/dashboard.html', context)


@login_required
def call_log_list(request):
    """List view for call logs"""
    call_logs = CallLog.objects.select_related('employee_id').order_by('-call_date')
    
    # Search functionality
    search = request.GET.get('search')
    if search:
        call_logs = call_logs.filter(
            Q(employee_id__employee_first_name__icontains=search) |
            Q(employee_id__employee_last_name__icontains=search)
        )
    
    # Date filtering
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    if date_from:
        call_logs = call_logs.filter(call_date__gte=date_from)
    if date_to:
        call_logs = call_logs.filter(call_date__lte=date_to)
    
    # Pagination
    paginator = Paginator(call_logs, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search': search,
        'date_from': date_from,
        'date_to': date_to,
    }
    
    return render(request, 'call_attendance/call_log_list.html', context)


@login_required
def call_log_create(request):
    """Create new call log entry"""
    if request.method == 'POST':
        form = CallLogForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, _("Call log created successfully"))
            return redirect('call_log_list')
    else:
        form = CallLogForm()
    
    return render(request, 'call_attendance/call_log_form.html', {'form': form})


@login_required
def call_log_edit(request, pk):
    """Edit existing call log"""
    call_log = get_object_or_404(CallLog, pk=pk)
    
    if request.method == 'POST':
        form = CallLogForm(request.POST, instance=call_log)
        if form.is_valid():
            form.save()
            messages.success(request, _("Call log updated successfully"))
            return redirect('call_log_list')
    else:
        form = CallLogForm(instance=call_log)
    
    return render(request, 'call_attendance/call_log_form.html', {
        'form': form, 'call_log': call_log
    })


@login_required
def call_log_bulk_upload(request):
    """Bulk upload call logs from CSV"""
    if request.method == 'POST':
        form = CallLogBulkUploadForm(request.POST, request.FILES)
        if form.is_valid():
            csv_file = request.FILES['csv_file']
            
            try:
                # Read CSV data
                decoded_file = csv_file.read().decode('utf-8')
                csv_data = csv.DictReader(io.StringIO(decoded_file))
                
                call_data = []
                for row in csv_data:
                    try:
                        call_data.append({
                            'employee_id': int(row['employee_id']),
                            'call_date': datetime.strptime(row['call_date'], '%Y-%m-%d').date(),
                            'call_duration_minutes': int(row['call_duration_minutes']),
                            'call_count': int(row.get('call_count', 0)),
                            'source': 'BULK_UPLOAD'
                        })
                    except (ValueError, KeyError) as e:
                        messages.error(request, f"Error in row: {row}. {str(e)}")
                        continue
                
                # Bulk create
                created, updated = CallLogService.bulk_create_call_logs(call_data)
                messages.success(request, 
                    f"Successfully processed {created} new and {updated} updated call logs")
                
                return redirect('call_log_list')
                
            except Exception as e:
                messages.error(request, f"Error processing file: {str(e)}")
    else:
        form = CallLogBulkUploadForm()
    
    return render(request, 'call_attendance/bulk_upload.html', {'form': form})


@login_required
def attendance_config(request):
    """Manage attendance configuration"""
    config = CallAttendanceService.get_active_config()
    
    if request.method == 'POST':
        if config:
            form = CallAttendanceConfigForm(request.POST, instance=config)
        else:
            form = CallAttendanceConfigForm(request.POST)
        
        if form.is_valid():
            form.save()
            messages.success(request, _("Configuration updated successfully"))
            return redirect('attendance_config')
    else:
        form = CallAttendanceConfigForm(instance=config)
    
    return render(request, 'call_attendance/config.html', {'form': form, 'config': config})


@login_required
def attendance_list(request):
    """List view for call attendances"""
    attendances = CallAttendance.objects.select_related(
        'employee_id', 'updated_by'
    ).order_by('-attendance_date')
    
    # Apply filters
    form = AttendanceFilterForm(request.GET)
    if form.is_valid():
        if form.cleaned_data.get('employee'):
            attendances = attendances.filter(employee_id=form.cleaned_data['employee'])
        if form.cleaned_data.get('start_date'):
            attendances = attendances.filter(attendance_date__gte=form.cleaned_data['start_date'])
        if form.cleaned_data.get('end_date'):
            attendances = attendances.filter(attendance_date__lte=form.cleaned_data['end_date'])
        if form.cleaned_data.get('status'):
            attendances = attendances.filter(attendance_status=form.cleaned_data['status'])
        if form.cleaned_data.get('source'):
            attendances = attendances.filter(source=form.cleaned_data['source'])
    
    # Pagination
    paginator = Paginator(attendances, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'filter_form': form,
    }
    
    return render(request, 'call_attendance/attendance_list.html', context)


@login_required
def manual_update_attendance(request):
    """Manual attendance update by TL/Manager with comprehensive validation"""
    validator = ComprehensiveValidator(request)
    
    if request.method == 'POST':
        form = ManualAttendanceUpdateForm(request.POST)
        if form.is_valid():
            try:
                # Prepare validation data
                validation_data = {
                    'employee_id': form.cleaned_data['employee_id'].id,
                    'attendance_date': form.cleaned_data['attendance_date'],
                    'call_duration_minutes': form.cleaned_data['call_duration_minutes'],
                    'call_count': form.cleaned_data['call_count'],
                    'reason': request.POST.get('reason', ''),
                    'loan_amount': form.cleaned_data.get('loan_amount', 0)
                }
                
                # Comprehensive validation
                validator.validate_attendance_update(validation_data, request.user)
                
                # Get old values for audit
                old_attendance = CallAttendance.objects.filter(
                    employee_id=form.cleaned_data['employee_id'],
                    attendance_date=form.cleaned_data['attendance_date']
                ).first()
                
                old_values = {}
                if old_attendance:
                    old_values = {
                        'attendance_status': old_attendance.attendance_status,
                        'call_duration_minutes': old_attendance.call_duration_minutes,
                        'call_count': old_attendance.call_count,
                        'loan_amount': getattr(old_attendance, 'loan_amount', 0)
                    }
                
                # Calculate attendance status based on call duration
                calculated_status = AttendanceValidator.calculate_attendance_status(
                    validation_data['call_duration_minutes']
                )
                
                # Create or update attendance
                attendance, created = CallAttendance.objects.get_or_create(
                    employee_id=form.cleaned_data['employee_id'],
                    attendance_date=form.cleaned_data['attendance_date'],
                    defaults={
                        'attendance_status': calculated_status,
                        'call_duration_minutes': validation_data['call_duration_minutes'],
                        'call_count': validation_data['call_count'],
                        'loan_amount': validation_data['loan_amount'],
                        'source': 'MANUAL',
                        'updated_by': request.user
                    }
                )
                
                if not created:
                    attendance.attendance_status = calculated_status
                    attendance.call_duration_minutes = validation_data['call_duration_minutes']
                    attendance.call_count = validation_data['call_count']
                    attendance.loan_amount = validation_data['loan_amount']
                    attendance.source = 'MANUAL'
                    attendance.updated_by = request.user
                    attendance.save()
                
                # Create audit log
                new_values = {
                    'attendance_status': attendance.attendance_status,
                    'call_duration_minutes': attendance.call_duration_minutes,
                    'call_count': attendance.call_count,
                    'loan_amount': attendance.loan_amount
                }
                
                AuditValidator.create_audit_log(
                    attendance, request.user, 'MANUAL_UPDATE',
                    old_values, new_values, validation_data['reason']
                )
                
                messages.success(request, _("Attendance updated successfully"))
                return redirect('call_attendance:attendance_list')
                
            except SalesValidationError as e:
                for field, error_list in e.message_dict.items():
                    if isinstance(error_list, list):
                        for error in error_list:
                            messages.error(request, f"{field}: {error}")
                    else:
                        messages.error(request, f"{field}: {error_list}")
            except Exception as e:
                messages.error(request, str(e))
    else:
        form = ManualAttendanceUpdateForm()
    
    # Get team members based on user role
    from base.templatetags.basefilters import is_reportingmanager
    if is_reportingmanager(request.user):
        team_members = Employee.objects.filter(
            employee_work_info__reporting_manager_id=request.user.employee_get
        )
    else:
        team_members = Employee.objects.filter(id=request.user.employee_get.id)
    
    context = {
        'form': form,
        'team_members': team_members,
        'is_manager': is_reportingmanager(request.user),
    }
    
    return render(request, 'call_attendance/manual_update.html', context)


@login_required
def employee_attendance_view(request):
    """Employee view of their own attendance"""
    try:
        employee = request.user.employee_get
    except Employee.DoesNotExist:
        messages.error(request, _("User not associated with an employee"))
        return redirect('dashboard')
    
    # Get current month data by default
    today = date.today()
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if not start_date:
        start_date = date(today.year, today.month, 1)
    else:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    
    if not end_date:
        end_date = today
    else:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # Get attendance records
    attendances = CallAttendance.objects.filter(
        employee_id=employee,
        attendance_date__range=[start_date, end_date]
    ).order_by('-attendance_date')
    
    # Get summary
    summary = CallAttendanceService.get_employee_attendance_summary(
        employee, start_date, end_date
    )
    
    context = {
        'employee': employee,
        'attendances': attendances,
        'summary': summary,
        'start_date': start_date,
        'end_date': end_date,
    }
    
    return render(request, 'call_attendance/employee_view.html', context)


@login_required
def audit_trail(request):
    """View audit trail for attendance updates"""
    # Check if user can view audit trail (admin or manager)
    if not (request.user.is_superuser or CallAttendanceService.can_update_attendance(request.user)):
        raise PermissionDenied(_("You don't have permission to view audit trail"))
    
    audits = CallAttendanceAudit.objects.select_related(
        'employee_id', 'updated_by'
    ).order_by('-timestamp')
    
    # Filtering
    employee_id = request.GET.get('employee')
    if employee_id:
        audits = audits.filter(employee_id=employee_id)
    
    start_date = request.GET.get('start_date')
    if start_date:
        audits = audits.filter(attendance_date__gte=start_date)
    
    end_date = request.GET.get('end_date')
    if end_date:
        audits = audits.filter(attendance_date__lte=end_date)
    
    # Pagination
    paginator = Paginator(audits, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'employees': Employee.objects.filter(is_active=True),
        'selected_employee': employee_id,
        'start_date': start_date,
        'end_date': end_date,
    }
    
    return render(request, 'call_attendance/audit_trail.html', context)


@login_required
@require_http_methods(["POST"])
def calculate_attendance(request):
    """AJAX endpoint to trigger attendance calculation"""
    if not request.user.is_superuser:
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    target_date = request.POST.get('date')
    if target_date:
        target_date = datetime.strptime(target_date, '%Y-%m-%d').date()
    else:
        target_date = date.today()
    
    try:
        result = CallAttendanceService.calculate_daily_attendance(target_date)
        return JsonResponse({
            'success': True,
            'message': f"Processed {result['processed']} records, skipped {result['skipped']}",
            'data': result
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def attendance_report(request):
    """Generate attendance report"""
    # Date range
    end_date = date.today()
    start_date = end_date - timedelta(days=30)  # Last 30 days
    
    if request.GET.get('start_date'):
        start_date = datetime.strptime(request.GET['start_date'], '%Y-%m-%d').date()
    if request.GET.get('end_date'):
        end_date = datetime.strptime(request.GET['end_date'], '%Y-%m-%d').date()
    
    # Get attendance data
    attendances = CallAttendance.objects.filter(
        attendance_date__range=[start_date, end_date]
    ).select_related('employee_id')
    
    # Generate summary statistics
    summary_stats = attendances.aggregate(
        total_records=Count('id'),
        total_present=Count('id', filter=Q(attendance_status='PRESENT')),
        total_half_day=Count('id', filter=Q(attendance_status='HALF_DAY')),
        total_absent=Count('id', filter=Q(attendance_status='ABSENT')),
        total_manual=Count('id', filter=Q(source='MANUAL')),
        total_call_minutes=Sum('call_duration_minutes'),
    )
    
    # Employee-wise summary
    employee_summary = {}
    for attendance in attendances:
        emp_id = attendance.employee_id.id
        if emp_id not in employee_summary:
            employee_summary[emp_id] = {
                'employee': attendance.employee_id,
                'present': 0,
                'half_day': 0,
                'absent': 0,
                'total_minutes': 0,
                'manual_updates': 0,
            }
        
        if attendance.attendance_status == 'PRESENT':
            employee_summary[emp_id]['present'] += 1
        elif attendance.attendance_status == 'HALF_DAY':
            employee_summary[emp_id]['half_day'] += 1
        else:
            employee_summary[emp_id]['absent'] += 1
        
        employee_summary[emp_id]['total_minutes'] += attendance.call_duration_minutes
        if attendance.source == 'MANUAL':
            employee_summary[emp_id]['manual_updates'] += 1
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'summary_stats': summary_stats,
        'employee_summary': employee_summary.values(),
        'attendances': attendances[:100],  # Limit for display
    }
    
    return render(request, 'call_attendance/report.html', context)


@login_required
def download_attendance_template(request):
    """Download Excel template for bulk attendance update"""
    user = request.user
    
    # Get team members
    from base.templatetags.basefilters import is_reportingmanager
    if is_reportingmanager(user):
        employees = Employee.objects.filter(
            employee_work_info__reporting_manager_id=user.employee_get
        )
    else:
        employees = Employee.objects.filter(id=user.employee_get.id)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Call Attendance Template"
    
    # Headers
    headers = [
        'Employee ID', 'Employee Name', 'Date', 'Status', 
        'Call Duration (Minutes)', 'Call Count', 'Leads Generated', 'Disbursements', 'Loan Amount',
        'Target Calls', 'Target Leads', 'Target Disbursements'
    ]
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Add employee data
    row = 2
    today = date.today()
    
    for emp in employees:
        ws.cell(row=row, column=1, value=emp.badge_id or emp.id)
        ws.cell(row=row, column=2, value=f"{emp.employee_first_name} {emp.employee_last_name}")
        ws.cell(row=row, column=3, value=today.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=4, value="PRESENT")
        ws.cell(row=row, column=5, value=0)  # Call Duration
        ws.cell(row=row, column=6, value=0)  # Call Count
        ws.cell(row=row, column=7, value=0)  # Leads Generated
        ws.cell(row=row, column=8, value=0)  # Disbursements
        ws.cell(row=row, column=9, value=0)  # Loan Amount
        ws.cell(row=row, column=10, value=50)  # Target Calls
        ws.cell(row=row, column=11, value=10)  # Target Leads
        ws.cell(row=row, column=12, value=2)  # Target Disbursements
        row += 1
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="call_attendance_template_{today.strftime("%Y%m%d")}.xlsx"'
    
    wb.save(response)
    return response


@login_required
@require_http_methods(["POST"])
def upload_bulk_attendance(request):
    """Upload bulk attendance data with comprehensive validation"""
    validator = ComprehensiveValidator(request)
    
    if 'attendance_file' not in request.FILES:
        return JsonResponse({'success': False, 'message': 'No file selected'})
    
    file = request.FILES['attendance_file']
    
    try:
        df = pd.read_excel(file)
        
        required_columns = ['Employee ID', 'Date', 'Call Duration (Minutes)', 'Call Count']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return JsonResponse({
                'success': False, 
                'message': f'Missing columns: {", ".join(missing_columns)}'
            })
        
        success_count = 0
        error_count = 0
        validation_errors = []
        
        for index, row in df.iterrows():
            try:
                employee = Employee.objects.filter(badge_id=row['Employee ID']).first()
                if not employee:
                    employee = Employee.objects.filter(id=row['Employee ID']).first()
                
                if not employee:
                    validation_errors.append(f"Row {index + 2}: Employee not found")
                    error_count += 1
                    continue
                
                attendance_date = pd.to_datetime(row['Date']).date()
                
                validation_data = {
                    'employee_id': employee.id,
                    'attendance_date': attendance_date,
                    'call_duration_minutes': int(row.get('Call Duration (Minutes)', 0)),
                    'call_count': int(row.get('Call Count', 0)),
                    'loan_amount': float(row.get('Loan Amount', 0)),
                    'reason': f'Bulk upload on {datetime.now().strftime("%Y-%m-%d %H:%M")}'
                }
                
                try:
                    validator.validate_attendance_update(validation_data, request.user)
                except SalesValidationError as ve:
                    validation_errors.append(f"Row {index + 2}: {ve}")
                    error_count += 1
                    continue
                
                calculated_status = AttendanceValidator.calculate_attendance_status(
                    validation_data['call_duration_minutes']
                )
                
                call_attendance, created = CallAttendance.objects.get_or_create(
                    employee_id=employee,
                    attendance_date=attendance_date,
                    defaults={
                        'attendance_status': calculated_status,
                        'call_duration_minutes': validation_data['call_duration_minutes'],
                        'call_count': validation_data['call_count'],
                        'leads_generated': int(row.get('Leads Generated', 0)),
                        'disbursements': int(row.get('Disbursements', 0)),
                        'loan_amount': validation_data['loan_amount'],
                        'source': 'BULK_UPLOAD',
                        'updated_by': request.user
                    }
                )
                
                if not created:
                    call_attendance.attendance_status = calculated_status
                    call_attendance.call_duration_minutes = validation_data['call_duration_minutes']
                    call_attendance.call_count = validation_data['call_count']
                    call_attendance.leads_generated = int(row.get('Leads Generated', 0))
                    call_attendance.disbursements = int(row.get('Disbursements', 0))
                    call_attendance.loan_amount = validation_data['loan_amount']
                    call_attendance.source = 'BULK_UPLOAD'
                    call_attendance.updated_by = request.user
                    call_attendance.save()
                
                success_count += 1
                
            except Exception as e:
                validation_errors.append(f"Row {index + 2}: {str(e)}")
                error_count += 1
                continue
        
        message = f'Processed {success_count} records successfully'
        if error_count > 0:
            message += f', {error_count} failed'
        
        return JsonResponse({
            'success': True, 
            'message': message,
            'details': {
                'success_count': success_count,
                'error_count': error_count,
                'validation_errors': validation_errors[:10]
            }
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})