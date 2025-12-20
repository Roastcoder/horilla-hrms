from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_http_methods
from django.contrib import messages
from django.utils.translation import gettext as _
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import pandas as pd
from datetime import datetime, date
from employee.models import Employee
from attendance.models import CallAttendance
from base.templatetags.basefilters import is_reportingmanager
from .models import IncentiveSlab, LoanType, Lead


@login_required
def dashboard(request):
    stats = {
        'total_leads': Lead.objects.count(),
        'disbursed_leads': Lead.objects.filter(status='DISBURSED').count(),
        'total_slabs': IncentiveSlab.objects.filter(is_active=True).count(),
        'loan_types': LoanType.objects.filter(is_active=True).count(),
    }
    
    return render(request, 'sales_incentive/dashboard.html', {'stats': stats})

@login_required
def incentive_slab_view(request):
    slabs = IncentiveSlab.objects.filter(is_active=True).order_by('min_amount')
    return render(request, 'sales_incentive/incentive_slabs.html', {'slabs': slabs})

@login_required
@require_http_methods(["GET", "POST"])
def incentive_slab_create(request):
    if request.method == "POST":
        try:
            min_amount = request.POST.get('min_amount')
            max_amount = request.POST.get('max_amount')
            incentive_per_lac = request.POST.get('incentive_per_lac')
            
            if not min_amount or not incentive_per_lac:
                messages.error(request, _("Min amount and incentive per lac are required."))
                return JsonResponse({'success': False})
            
            IncentiveSlab.objects.create(
                min_amount=min_amount,
                max_amount=max_amount if max_amount else None,
                incentive_per_lac=incentive_per_lac
            )
            
            messages.success(request, _("Incentive slab created successfully."))
            return JsonResponse({'success': True})
            
        except Exception as e:
            messages.error(request, _(f"Error creating slab: {str(e)}"))
            return JsonResponse({'success': False})
    
    return render(request, 'sales_incentive/slab_form.html')

@login_required
@require_http_methods(["GET", "POST"])
def incentive_slab_update(request, slab_id):
    slab = IncentiveSlab.objects.get(id=slab_id)
    
    if request.method == "POST":
        try:
            slab.min_amount = request.POST.get('min_amount')
            slab.max_amount = request.POST.get('max_amount') if request.POST.get('max_amount') else None
            slab.incentive_per_lac = request.POST.get('incentive_per_lac')
            slab.save()
            
            messages.success(request, _("Incentive slab updated successfully."))
            return JsonResponse({'success': True})
            
        except Exception as e:
            messages.error(request, _(f"Error updating slab: {str(e)}"))
            return JsonResponse({'success': False})
    
    return render(request, 'sales_incentive/slab_form.html', {'slab': slab})

@login_required
@require_http_methods(["POST"])
def incentive_slab_delete(request, slab_id):
    try:
        slab = IncentiveSlab.objects.get(id=slab_id)
        slab.is_active = False
        slab.save()
        
        messages.success(request, _("Incentive slab deleted successfully."))
        return JsonResponse({'success': True})
        
    except Exception as e:
        messages.error(request, _(f"Error deleting slab: {str(e)}"))
        return JsonResponse({'success': False})

@login_required
def loan_type_view(request):
    loan_types = LoanType.objects.filter(is_active=True).order_by('name')
    return render(request, 'sales_incentive/loan_types.html', {'loan_types': loan_types})

@login_required
@require_http_methods(["GET", "POST"])
def loan_type_create(request):
    if request.method == "POST":
        try:
            name = request.POST.get('name')
            points_per_lac = request.POST.get('points_per_lac')
            
            if not name or not points_per_lac:
                messages.error(request, _("Name and points per lac are required."))
                return JsonResponse({'success': False})
            
            LoanType.objects.create(
                name=name,
                points_per_lac=points_per_lac
            )
            
            messages.success(request, _("Loan type created successfully."))
            return JsonResponse({'success': True})
            
        except Exception as e:
            messages.error(request, _(f"Error creating loan type: {str(e)}"))
            return JsonResponse({'success': False})
    
    return render(request, 'sales_incentive/loan_type_form.html')

@login_required
@require_http_methods(["GET", "POST"])
def loan_type_update(request, type_id):
    loan_type = LoanType.objects.get(id=type_id)
    
    if request.method == "POST":
        try:
            loan_type.name = request.POST.get('name')
            loan_type.points_per_lac = request.POST.get('points_per_lac')
            loan_type.save()
            
            messages.success(request, _("Loan type updated successfully."))
            return JsonResponse({'success': True})
            
        except Exception as e:
            messages.error(request, _(f"Error updating loan type: {str(e)}"))
            return JsonResponse({'success': False})
    
    return render(request, 'sales_incentive/loan_type_form.html', {'loan_type': loan_type})

@login_required
@require_http_methods(["POST"])
def loan_type_delete(request, type_id):
    try:
        loan_type = LoanType.objects.get(id=type_id)
        loan_type.is_active = False
        loan_type.save()
        
        messages.success(request, _("Loan type deleted successfully."))
        return JsonResponse({'success': True})
        
    except Exception as e:
        messages.error(request, _(f"Error deleting loan type: {str(e)}"))
        return JsonResponse({'success': False})

@login_required
def lead_view(request):
    leads = Lead.objects.all().order_by('-created_date')
    employees = Employee.objects.filter(is_active=True)
    loan_types = LoanType.objects.filter(is_active=True)
    return render(request, 'sales_incentive/leads.html', {
        'leads': leads,
        'employees': employees,
        'loan_types': loan_types
    })

@login_required
@require_http_methods(["GET", "POST"])
def lead_create(request):
    if request.method == "POST":
        try:
            employee_id = request.POST.get('employee_id')
            loan_type_id = request.POST.get('loan_type_id')
            loan_amount = request.POST.get('loan_amount')
            status = request.POST.get('status', 'NEW')
            
            if not employee_id or not loan_type_id or not loan_amount:
                messages.error(request, _("Employee, loan type and amount are required."))
                return JsonResponse({'success': False})
            
            Lead.objects.create(
                employee_id=employee_id,
                loan_type_id=loan_type_id,
                loan_amount=loan_amount,
                status=status
            )
            
            messages.success(request, _("Lead created successfully."))
            return JsonResponse({'success': True})
            
        except Exception as e:
            messages.error(request, _(f"Error creating lead: {str(e)}"))
            return JsonResponse({'success': False})
    
    employees = Employee.objects.filter(is_active=True)
    loan_types = LoanType.objects.filter(is_active=True)
    return render(request, 'sales_incentive/lead_form.html', {
        'employees': employees,
        'loan_types': loan_types
    })

@login_required
@require_http_methods(["GET", "POST"])
def lead_update(request, lead_id):
    lead = Lead.objects.get(id=lead_id)
    
    if request.method == "POST":
        try:
            lead.employee_id = request.POST.get('employee_id')
            lead.loan_type_id = request.POST.get('loan_type_id')
            lead.loan_amount = request.POST.get('loan_amount')
            lead.status = request.POST.get('status')
            
            if lead.status == 'DISBURSED' and not lead.disbursed_date:
                lead.disbursed_date = timezone.now()
            
            lead.save()
            
            messages.success(request, _("Lead updated successfully."))
            return JsonResponse({'success': True})
            
        except Exception as e:
            messages.error(request, _(f"Error updating lead: {str(e)}"))
            return JsonResponse({'success': False})
    
    employees = Employee.objects.filter(is_active=True)
    loan_types = LoanType.objects.filter(is_active=True)
    return render(request, 'sales_incentive/lead_form.html', {
        'lead': lead,
        'employees': employees,
        'loan_types': loan_types
    })

@login_required
@require_http_methods(["POST"])
def lead_delete(request, lead_id):
    try:
        lead = Lead.objects.get(id=lead_id)
        lead.delete()
        
        messages.success(request, _("Lead deleted successfully."))
        return JsonResponse({'success': True})
        
    except Exception as e:
        messages.error(request, _(f"Error deleting lead: {str(e)}"))
        return JsonResponse({'success': False})

@login_required
def manual_attendance(request):
    user = request.user
    
    # Get team members based on user role
    if is_reportingmanager(user):
        team_members = Employee.objects.filter(
            employee_work_info__reporting_manager_id=user.employee_get
        )
    else:
        team_members = Employee.objects.filter(id=user.employee_get.id)
    
    context = {
        'team_members': team_members,
        'is_manager': is_reportingmanager(user),
    }
    return render(request, 'sales_incentive/manual_attendance.html', context)

@login_required
def download_template(request):
    user = request.user
    
    # Get team members
    if is_reportingmanager(user):
        employees = Employee.objects.filter(
            employee_work_info__reporting_manager_id=user.employee_get
        )
    else:
        employees = Employee.objects.filter(id=user.employee_get.id)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Template"
    
    # Headers
    headers = [
        'Employee ID', 'Employee Name', 'Date', 'Status', 
        'Calls Made', 'Leads Generated', 'Disbursements', 
        'Target Calls', 'Target Leads', 'Target Disbursements'
    ]
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Add employee data
    row = 2
    today = date.today()
    
    for emp in employees:
        ws.cell(row=row, column=1, value=emp.badge_id or emp.id)
        ws.cell(row=row, column=2, value=f"{emp.employee_first_name} {emp.employee_last_name}")
        ws.cell(row=row, column=3, value=today.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=4, value="PRESENT")
        ws.cell(row=row, column=5, value=0)
        ws.cell(row=row, column=6, value=0)
        ws.cell(row=row, column=7, value=0)
        ws.cell(row=row, column=8, value=50)
        ws.cell(row=row, column=9, value=10)
        ws.cell(row=row, column=10, value=2)
        row += 1
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="attendance_template_{today.strftime("%Y%m%d")}.xlsx"'
    
    wb.save(response)
    return response

@login_required
@require_http_methods(["POST"])
def upload_attendance(request):
    if 'attendance_file' not in request.FILES:
        messages.error(request, _("Please select a file to upload."))
        return JsonResponse({'success': False, 'message': 'No file selected'})
    
    file = request.FILES['attendance_file']
    
    try:
        # Read Excel file
        df = pd.read_excel(file)
        
        # Validate columns
        required_columns = [
            'Employee ID', 'Date', 'Status', 'Calls Made', 
            'Leads Generated', 'Disbursements'
        ]
        
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            messages.error(request, _(f"Missing columns: {', '.join(missing_columns)}"))
            return JsonResponse({'success': False, 'message': f'Missing columns: {", ".join(missing_columns)}'})
        
        # Process each row
        success_count = 0
        error_count = 0
        
        for index, row in df.iterrows():
            try:
                # Find employee
                employee = Employee.objects.filter(
                    badge_id=row['Employee ID']
                ).first()
                
                if not employee:
                    employee = Employee.objects.filter(
                        id=row['Employee ID']
                    ).first()
                
                if not employee:
                    error_count += 1
                    continue
                
                # Parse date
                attendance_date = pd.to_datetime(row['Date']).date()
                
                # Create or update CallAttendance
                call_attendance, created = CallAttendance.objects.get_or_create(
                    employee_id=employee,
                    date=attendance_date,
                    defaults={
                        'status': row['Status'],
                        'calls_made': int(row.get('Calls Made', 0)),
                        'leads_generated': int(row.get('Leads Generated', 0)),
                        'disbursements': int(row.get('Disbursements', 0)),
                        'source': 'MANUAL'
                    }
                )
                
                if not created:
                    # Update existing record
                    call_attendance.status = row['Status']
                    call_attendance.calls_made = int(row.get('Calls Made', 0))
                    call_attendance.leads_generated = int(row.get('Leads Generated', 0))
                    call_attendance.disbursements = int(row.get('Disbursements', 0))
                    call_attendance.source = 'MANUAL'
                    call_attendance.save()
                
                success_count += 1
                
            except Exception as e:
                error_count += 1
                continue
        
        if success_count > 0:
            messages.success(request, _(f"Successfully updated {success_count} attendance records."))
        
        if error_count > 0:
            messages.warning(request, _(f"{error_count} records failed to update."))
        
        return JsonResponse({
            'success': True, 
            'message': f'Processed {success_count} records successfully, {error_count} failed'
        })
        
    except Exception as e:
        messages.error(request, _(f"Error processing file: {str(e)}"))
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})